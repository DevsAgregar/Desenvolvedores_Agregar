import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

apiUrl = 'http://mhzenergiasolar.api.groner.app'
token_apikey = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1bmlxdWVfbmFtZSI6WyI3MiIsIkFHUkVHQVIiXSwianRpIjoiMTc2NDg3MjAxMWQ0NDA0YmEwNzI1NzU3MGVhZGNiMDIiLCJyb2xlIjoiQWRtaW5pc3RyYWRvciIsIkZ1bmNhbyI6IkRpcmV0b3IgQ29tZXJjaWFsIiwiVGltZVpvbmUiOiJBbWVyaWNhL1Nhb19QYXVsbyIsIkVtYWlsIjoiYWdyZWdhckBtaHplbmVyZ2lhc29sYXIuY29tLmJyIiwiSWQiOiI3MiIsIm5iZiI6MTczNDM3NzQ3OCwiZXhwIjoxNzc3NTc3NDc4LCJpYXQiOjE3MzQzNzc0NzgsImlzcyI6InJvbmFsZG8iLCJhdWQiOiJ0b3B6ZXJhIn0.MXa57uqxF6vDR4MpYsUDqpGK18X6Yog3d5vWEY8PyGs'
headers = {
    'Content-Type': 'application/json',
    'accept': 'application/json',
    'Authorization': f'Bearer {token_apikey}'
}

response = requests.get(
    f'{apiUrl}/api/Agendamento',
    headers=headers
)

# Parse the JSON response
data = response.json()

# Convert the JSON data to a DataFrame
df = pd.DataFrame(data)

# Save the DataFrame to an Excel file
df.to_excel('Agendamentos.xlsx', index=False)

# Load the workbook and select the active worksheet
wb = load_workbook('Agendamentos.xlsx')
ws = wb.active

# Define styles for the header
header_font = Font(bold=True)
header_fill = PatternFill(start_color='FFFF00',
                          end_color='FFFF00', fill_type='solid')

# Apply styles to the header row
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# Adjust column width automatically
for column in ws.columns:
    max_length = 0
    column = list(column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Save the formatted workbook
wb.save('Agendamentos_formatado.xlsx')

print("Dados foram salvos no arquivo Agendamentos_formatado.xlsx")
