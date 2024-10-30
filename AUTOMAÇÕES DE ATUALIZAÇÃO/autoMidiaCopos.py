import os
import shutil
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook
import sys

def executar_script():
    max_tentativas = 2
    tentativas = 0
    
    while tentativas < max_tentativas:
        navegador = None
        try:
            # Diretório de downloads
            download_dir = "C:\\Users\\User\\Downloads"

            # Configurações do Chrome
            chrome_options = Options()
            chrome_options.add_experimental_option("detach", True)
            chrome_options.add_experimental_option("prefs", {
                "download.default_directory": download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True
            })
            chrome_options.add_argument('--start-maximized')
            
            # Instala o ChromeDriver
            chrome_install = ChromeDriverManager().install()

            # Caminho do chromedriver
            folder = os.path.dirname(chrome_install)
            chromedriver_path = os.path.join(folder, "chromedriver.exe")
            service = Service(chromedriver_path)

            # Inicializa o navegador
            navegador = webdriver.Chrome(service=service, options=chrome_options)
            
            def baixar_cadastros():
                # Entra no Gestão Click
                navegador.get('https://gestaoclick.com.br/login/')
                time.sleep(2)
                
                # Faz login no Gestão Click
                navegador.find_element(By.XPATH, '//*[@id="UsuarioEmail"]').send_keys('financeiro@midiacopos.com.br')
                time.sleep(1)
                navegador.find_element(By.XPATH, '//*[@id="UsuarioSenha"]').send_keys('Qawsed54#')
                time.sleep(1)
                navegador.find_element(By.XPATH, '//*[@id="login"]/div[5]/button').click()
                time.sleep(10)
            
                # Acessa os cadastros de produtos
                navegador.get('https://gestaoclick.com/relatorios_cadastros/relatorio_produtos')
                time.sleep(5)

                # Gera o relatório
                navegador.find_element(By.XPATH, '/html/body/div[2]/div/div/aside[2]/div/section[2]/div/div/div/form/div[2]/button[1]').click()
                time.sleep(15)

                # Muda o foco para a nova aba
                navegador.switch_to.window(navegador.window_handles[-1])

                # Exporta o relatório
                WebDriverWait(navegador, 40).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="dv"]/caption/button'))).click()
                time.sleep(5)

                # Caminho do arquivo baixado
                downloaded_file = os.path.join(download_dir, "relatorio_produtos.xlsx")
                
                # Caminho do arquivo de origem (onde o arquivo baixado será movido)
                source_file = 'G:\\Drives compartilhados\\Agregar Negócios - Drive Geral\\Agregar Clientes Ativos\\MIDIA COPOS\\3. Finanças\\3 - Relatórios Financeiros\\CADASTROS DE PRODUTOS (SISTEMA).xlsx'
            
                # Move o arquivo baixado para o diretório de origem, substituindo o arquivo existente
                if os.path.exists(downloaded_file):
                    shutil.move(downloaded_file, source_file)
                    
                # Caminho do arquivo de destino (onde os dados serão copiados)
                destination_file = 'G:\\Drives compartilhados\\Agregar Negócios - Drive Geral\\Agregar Clientes Ativos\\MIDIA COPOS\\3. Finanças\\3 - Relatórios Financeiros\\PRECIFICAÇÃO - versão POWERBI.xlsm'
                
                # Copia os dados da planilha de origem para a planilha de destino
                copiar_dados(source_file, destination_file)
                    
            def copiar_dados(source_file, destination_file):
                # Abre a planilha de origem
                wb_source = load_workbook(source_file)
                ws_source = wb_source.active

                # Abre a planilha de destino
                wb_dest = load_workbook(destination_file, keep_vba=True)
                ws_dest = wb_dest['BANCO DE DADOS PRODUTOS']

                # Define a linha inicial para colar os dados, começando na linha 2
                start_row = 2

                # Copia os dados das colunas A, B e E a partir da linha 3 da planilha de origem
                for row_idx, row in enumerate(ws_source.iter_rows(min_row=3, max_col=5), start=start_row):
                    # Coloca o valor da coluna A na coluna A da planilha de destino
                    ws_dest.cell(row=row_idx, column=1, value=row[0].value)  # Coluna A

                    # Coloca o valor da coluna B na coluna B da planilha de destino
                    ws_dest.cell(row=row_idx, column=2, value=row[1].value)  # Coluna B

                    # Coloca o valor da coluna E na coluna C da planilha de destino
                    ws_dest.cell(row=row_idx, column=3, value=row[4].value)  # Coluna E (origem) para Coluna C (destino)

                # Salva a planilha de destino
                wb_dest.save(destination_file)

            baixar_cadastros()
            
            return 0 # para sucesso
            
        except Exception as e:
            tentativas += 1
            if tentativas < max_tentativas:
                time.sleep(5)
            else:
                return 1 # para falha
            
        finally:
            if navegador:
                navegador.quit()
    return 1

sys.exit(executar_script())